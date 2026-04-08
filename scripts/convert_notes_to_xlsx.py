#!/usr/bin/env python3
"""Convert exported note JSON data to a styled Excel spreadsheet.

Pure data script — does not launch a browser.
"""

import argparse
import json
import sys
from pathlib import Path

from _bootstrap_env import ensure_local_venv

ensure_local_venv()

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


AVAILABLE_COLUMNS = [
    "note_id",
    "note_url",
    "title",
    "desc",
    "tags",
    "nickname",
    "note_type",
    "publish_time",
    "update_time",
    "ip_location",
    "liked_count",
    "collected_count",
    "comment_count",
    "share_count",
    "source_keyword",
]

DEFAULT_COLUMNS = "title,nickname,note_url,tags,desc,publish_time,liked_count"

MAX_COL_WIDTH = 50


# ---------------------------------------------------------------------------
# Data loading & normalisation
# ---------------------------------------------------------------------------


def _parse_tags(value: object) -> str:
    """Return a human-readable tag string like ``#tag1 #tag2``."""
    if isinstance(value, list):
        return " ".join(f"#{t}" for t in value if t)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return ""
        # May already be a pre-formatted string like "#tag1 #tag2"
        if stripped.startswith("#"):
            return stripped
        # Try to parse as JSON array
        try:
            parsed = json.loads(stripped)
            if isinstance(parsed, list):
                return " ".join(f"#{t}" for t in parsed if t)
        except (json.JSONDecodeError, TypeError):
            pass
        return stripped
    return str(value) if value else ""


def load_json(path: str) -> object:
    """Read and parse JSON from *path*."""
    text = Path(path).read_text(encoding="utf-8").strip()
    if not text:
        raise ValueError(f"Input file is empty: {path}")
    return json.loads(text)


def extract_rows(payload: object) -> list[dict]:
    """Auto-detect the data format and return a list of flat row dicts."""
    # Format 1: bare array of flat objects
    if isinstance(payload, list):
        return [_normalise_item(item) for item in payload]

    if not isinstance(payload, dict):
        raise ValueError("Input JSON must be an array or an object")

    items = payload.get("items")
    if not isinstance(items, list):
        raise ValueError("Object-style input must contain an 'items' array")

    return [_normalise_item(item) for item in items]


def _normalise_item(item: object) -> dict:
    """Turn a single item (flat or with ``article_info``) into a flat dict."""
    if not isinstance(item, dict):
        raise ValueError(f"Each item must be a JSON object, got {type(item).__name__}")

    # Format 2: item with article_info key (batch_generate_comment_materials output)
    if "article_info" in item:
        article = item["article_info"]
        if not isinstance(article, dict):
            raise ValueError("article_info must be a JSON object")
        row: dict = {}
        # Top-level fields first
        for key in ("note_id", "note_url", "source_keyword"):
            if key in item:
                row[key] = item[key]
        # Merge article_info fields
        row.update(article)
        return row

    # Format 3: already flat
    return dict(item)


def prepare_rows(
    raw_rows: list[dict],
    columns: list[str],
) -> list[dict]:
    """Select requested *columns* and format special fields."""
    prepared: list[dict] = []
    for raw in raw_rows:
        row: dict = {}
        for col in columns:
            value = raw.get(col, "")
            if col == "tags":
                value = _parse_tags(value)
            elif value is None:
                value = ""
            row[col] = str(value)
        prepared.append(row)
    return prepared


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------


def _estimate_col_width(values: list[str], header: str) -> float:
    """Return an estimated column width capped at *MAX_COL_WIDTH*."""
    max_len = len(header)
    for v in values:
        # Rough CJK awareness: wide chars count as ~2 columns
        length = 0
        for ch in v:
            length += 2 if ord(ch) > 0x7F else 1
        if length > max_len:
            max_len = length
    return min(max_len + 2, MAX_COL_WIDTH)


def write_xlsx(
    rows: list[dict],
    columns: list[str],
    output_path: str,
) -> None:
    """Write *rows* to a styled XLSX file at *output_path*."""
    wb = Workbook()
    ws = wb.active
    ws.title = "notes"

    # -- Styles ----------------------------------------------------------
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # -- Header row ------------------------------------------------------
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # -- Data rows -------------------------------------------------------
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            align_kwargs: dict = {"horizontal": "left", "vertical": "top"}
            if col_name == "desc":
                align_kwargs["wrap_text"] = True
            cell.alignment = Alignment(**align_kwargs)

    # -- Auto column width -----------------------------------------------
    for col_idx, col_name in enumerate(columns, start=1):
        values = [row.get(col_name, "") for row in rows]
        width = _estimate_col_width(values, col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # -- Freeze first row ------------------------------------------------
    ws.freeze_panes = "A2"

    # -- Save ------------------------------------------------------------
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert exported note JSON data to a styled Excel spreadsheet.",
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Input JSON file path.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Output XLSX file path.",
    )
    parser.add_argument(
        "--columns",
        default=DEFAULT_COLUMNS,
        help=(
            "Comma-separated list of columns to include. "
            f"Default: {DEFAULT_COLUMNS}"
        ),
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    # Validate column names early
    requested = [c.strip() for c in args.columns.split(",") if c.strip()]
    valid_set = set(AVAILABLE_COLUMNS)
    unknown = [c for c in requested if c not in valid_set]
    if unknown:
        parser.error(
            f"Unknown column(s): {', '.join(unknown)}. "
            f"Available: {', '.join(AVAILABLE_COLUMNS)}"
        )

    payload = load_json(args.input)
    raw_rows = extract_rows(payload)
    rows = prepare_rows(raw_rows, requested)
    write_xlsx(rows, requested, args.output)

    print(f"Wrote {len(rows)} row(s) to {args.output}")


if __name__ == "__main__":
    main()
