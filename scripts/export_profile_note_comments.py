import argparse
import json
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from _bootstrap_env import ensure_local_venv

ensure_local_venv()

from openpyxl import Workbook

from action_delay import resolve_delay_seconds, sleep_random_cooldown
from batch_generate_comment_materials import (
    build_article_info,
    fetch_note_data,
    parse_note_id_from_url,
)
from list_profile_notes import list_profile_notes
from note_content import clean_text
from official_risk_guard import OfficialRiskDetectedError, RiskStopTracker


WORKSPACE_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_OUTPUT_DIR = WORKSPACE_ROOT / "deliveries" / "rednote_profile_comment_sheets"

ARTICLES_COLUMNS = [
    "note_id",
    "title",
    "nickname",
    "note_url",
    "tags",
    "desc",
    "comment_1",
    "comment_2",
    "comment_3",
    "comment_4",
    "comment_5",
    "matched_topics",
]

COMMENTS_COLUMNS = [
    "note_id",
    "title",
    "nickname",
    "note_url",
    "comment_index",
    "comment_text",
]

FAILED_COLUMNS = [
    "note_id",
    "title",
    "note_url",
    "error",
]

RULED_COMMENT_POOLS = OrderedDict(
    [
        (("ai", "智能", "大模型"), ["像开外挂", "AI真能打", "效率真狠", "这也太快了", "工具感很强"]),
        (("标书", "编标"), ["标书人有救", "太懂标书人", "这条真对口", "刚需内容啊", "先存着用"]),
        (("投标", "招标"), ["投标人狂喜", "投标党必看", "太懂投标了", "这波太对口", "真能少熬夜"]),
        (("保姆级", "教学", "教程", "操作", "演示"), ["跟着就会了", "这篇真细", "新手真友好", "流程很顺手", "讲得挺清楚"]),
        (("升级", "更新", "重磅"), ["这次像升级", "这波有变化", "终于更完整", "企业版更稳", "这次真走心"]),
        (("避坑", "误区", "雷区", "低分"), ["这坑真要命", "这条能避雷", "评委就看这", "太容易踩了", "这点太关键"]),
        (("查重", "重复"), ["查重刚需", "这步太重要", "终于不撞稿", "这条真救命", "重复率有救"]),
        (("协作", "资料", "版本", "材料慢"), ["协作乱最伤", "版本地狱了", "这句太真了", "流程得救了", "资料终于顺"]),
        (("采购方式", "采购"), ["终于讲明白", "小白也能懂", "这条真入门", "一看就懂了", "这篇够直白"]),
        (("技术方案", "技术分"), ["技术分保命", "这点真关键", "方案党速看", "评委真看这", "低分太亏了"]),
        (("加班", "节后", "日常"), ["太懂打工人", "这班加麻了", "像我昨天", "隔屏都累", "真是日常啊"]),
        (("作业帮",), ["投标版作业帮", "这比模板强", "像开挂一样", "我先试试", "这也太香了"]),
        (("企业版", "团队"), ["企业版刚需", "团队版更香", "这波适合团队", "老板会心动", "像真生产力"]),
        (("10分钟", "十分钟", "百页"), ["十分钟真香", "百页也敢上", "这速度离谱", "看着就省命", "效率太猛了"]),
    ]
)

TITLE_PATTERN_COMMENTS = OrderedDict(
    [
        (("必看",), ["这条必码", "先码住了"]),
        (("速存", "速进"), ["先存再说", "这就点开"]),
        (("一文看懂",), ["终于讲透了", "这回懂了"]),
        (("快用这个",), ["我先试试", "这就上手"]),
        (("还不知道",), ["现在知道了", "刚好补课"]),
        (("请进",), ["这就进来了", "我先围观"]),
    ]
)

FALLBACK_COMMENTS = [
    "先收藏了",
    "这思路行",
    "有点东西",
    "很像实战",
    "这条挺真",
    "真能落地吗",
    "这角度行",
    "我先记下",
    "挺接地气",
    "这波实用",
]


def _normalize_text(article: dict) -> str:
    parts = [
        article.get("title", ""),
        " ".join(article.get("tags", [])),
        article.get("nickname", ""),
    ]
    return clean_text(" ".join(parts)).lower()


def _comment_length(comment: str) -> int:
    return len("".join(str(comment).split()))


def _append_unique(target: list[str], candidates: list[str]) -> None:
    existing = set(target)
    for candidate in candidates:
        candidate = clean_text(candidate)
        if not candidate or candidate in existing:
            continue
        if _comment_length(candidate) > 10:
            continue
        target.append(candidate)
        existing.add(candidate)


def generate_short_comments(article: dict) -> tuple[list[str], list[str]]:
    text = _normalize_text(article)
    desc_text = clean_text(article.get("desc", "")).lower()
    title = clean_text(article.get("title", ""))
    comments: list[str] = []
    matched_topics: list[str] = []
    specific_pools: list[list[str]] = []
    generic_pools: list[list[str]] = []

    for keywords, pool in RULED_COMMENT_POOLS.items():
        hit_in_primary = any(keyword.lower() in text for keyword in keywords)
        if hit_in_primary:
            matched_topics.extend(list(keywords))
            if keywords in (
                ("ai", "智能", "大模型"),
                ("标书", "编标"),
                ("投标", "招标"),
            ):
                generic_pools.append(pool)
            else:
                specific_pools.append(pool)

    if not specific_pools and not generic_pools:
        for keywords, pool in RULED_COMMENT_POOLS.items():
            if not any(keyword.lower() in desc_text for keyword in keywords):
                continue
            matched_topics.extend(list(keywords))
            if keywords in (
                ("ai", "智能", "大模型"),
                ("标书", "编标"),
                ("投标", "招标"),
            ):
                generic_pools.append(pool)
            else:
                specific_pools.append(pool)

    for patterns, pool in TITLE_PATTERN_COMMENTS.items():
        if any(pattern in title for pattern in patterns):
            specific_pools.append(pool)

    if "？" in title or "?" in title:
        specific_pools.append(["这问得真准", "我也想知道"])
    if "别再" in title:
        specific_pools.append(["这句太扎心", "真的别硬扛"])
    if "终于" in title or "完成" in article.get("desc", ""):
        specific_pools.append(["这下省心了", "终于等到了"])

    matched_pools = list(specific_pools)
    if len(matched_pools) < 2:
        matched_pools.extend(generic_pools[:2])
    elif len(comments) < 3:
        matched_pools.extend(generic_pools[:1])

    for candidate_index in range(5):
        made_progress = False
        for pool in matched_pools:
            if candidate_index >= len(pool):
                continue
            before_count = len(comments)
            _append_unique(comments, [pool[candidate_index]])
            if len(comments) > before_count:
                made_progress = True
            if len(comments) >= 5:
                break
        if len(comments) >= 5 or not made_progress:
            break

    _append_unique(comments, FALLBACK_COMMENTS)
    return comments[:5], list(OrderedDict.fromkeys(matched_topics))[:6]


def load_note_list(input_path: str | None) -> list[dict]:
    if not input_path:
        return []
    payload = json.loads(Path(input_path).read_text(encoding="utf-8"))
    notes = payload.get("notes", []) if isinstance(payload, dict) else []
    if not isinstance(notes, list):
        raise ValueError("输入 JSON 缺少 notes 数组")
    return [note for note in notes if isinstance(note, dict)]


def fetch_notes_from_profile(
    profile_url: str,
    *,
    limit: int | None,
    max_scroll_rounds: int,
    max_idle_scroll_rounds: int,
    delay_seconds: float,
) -> list[dict]:
    result = list_profile_notes(
        profile_url=profile_url,
        limit=limit,
        max_scroll_rounds=max_scroll_rounds,
        max_idle_scroll_rounds=max_idle_scroll_rounds,
        delay_seconds=delay_seconds,
    )
    notes = result.get("notes", []) if isinstance(result, dict) else []
    if not result.get("ok", False):
        raise RuntimeError(result.get("message") or result.get("reason") or "主页抓取失败")
    if not isinstance(notes, list) or not notes:
        raise RuntimeError("主页没有抓到可用笔记")
    return notes


def build_comment_records(
    notes: list[dict],
    delay_seconds: float = 0.0,
    risk_stop_after: int = 3,
    note_read_cooldown_min_seconds: float = 4.0,
    note_read_cooldown_max_seconds: float = 7.0,
    batch_read_cooldown_every: int = 3,
    batch_read_cooldown_min_seconds: float = 20.0,
    batch_read_cooldown_max_seconds: float = 35.0,
) -> tuple[list[dict], list[dict], dict]:
    if int(batch_read_cooldown_every) <= 0:
        raise ValueError("batch_read_cooldown_every must be greater than 0")

    records: list[dict] = []
    failed: list[dict] = []
    attempted_note_reads = 0
    risk_tracker = RiskStopTracker(stop_after=risk_stop_after)

    for index, note in enumerate(notes, start=1):
        if risk_tracker.stopped_due_to_risk:
            break

        note_url = clean_text(note.get("note_url", ""))
        if not note_url:
            continue
        note_id = clean_text(note.get("id", "")) or parse_note_id_from_url(note_url) or f"note_{index}"
        print(f"📘 读取第 {index}/{len(notes)} 篇：{note_id}")
        if attempted_note_reads > 0:
            sleep_random_cooldown(
                "读取下一篇文章前冷却",
                note_read_cooldown_min_seconds,
                note_read_cooldown_max_seconds,
            )

        attempted_note_reads += 1
        try:
            note_data = fetch_note_data(note_url, delay_seconds)
            risk_tracker.record_success()
            article = build_article_info(note_data)
            comments, matched_topics = generate_short_comments(article)
            while len(comments) < 5:
                comments.append(f"先看这条{len(comments) + 1}")

            records.append(
                {
                    "note_id": note_id,
                    "title": article.get("title", "") or clean_text(note.get("title", "")),
                    "nickname": article.get("nickname", "") or clean_text(note.get("nickname", "")),
                    "note_url": note_url,
                    "tags": article.get("tags", []),
                    "desc": article.get("desc", ""),
                    "comments": comments[:5],
                    "matched_topics": matched_topics,
                }
            )
        except OfficialRiskDetectedError as exc:
            risk_state = risk_tracker.record_risk(exc.matched_phrase, exc.detail)
            error_text = clean_text(str(exc)) or exc.__class__.__name__
            print(f"⚠️ 第 {index} 篇触发官方风控：{error_text}")
            failed.append(
                {
                    "note_id": note_id,
                    "title": clean_text(note.get("title", "")),
                    "note_url": note_url,
                    "error": error_text,
                }
            )
            if risk_state["stopped_due_to_risk"]:
                break
        except Exception as exc:
            error_text = clean_text(str(exc)) or exc.__class__.__name__
            print(f"⚠️ 第 {index} 篇失败：{error_text}")
            failed.append(
                {
                    "note_id": note_id,
                    "title": clean_text(note.get("title", "")),
                    "note_url": note_url,
                    "error": error_text,
                }
            )
        if (
            attempted_note_reads > 0
            and attempted_note_reads % int(batch_read_cooldown_every) == 0
            and not risk_tracker.stopped_due_to_risk
        ):
            sleep_random_cooldown(
                "正文抓取批次冷却",
                batch_read_cooldown_min_seconds,
                batch_read_cooldown_max_seconds,
            )

    summary = {
        "attempted_note_reads": attempted_note_reads,
        **risk_tracker.snapshot(),
    }
    return records, failed, summary


def write_outputs(
    records: list[dict],
    failed: list[dict],
    summary: dict,
    output_dir: Path,
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%dT%H%M%S")
    json_path = output_dir / f"{run_id}.json"
    xlsx_path = output_dir / f"{run_id}.xlsx"

    json_payload = {
        "run_id": run_id,
        "count": len(records),
        "failed_count": len(failed),
        "summary": summary,
        "items": records,
        "failed_items": failed,
    }
    json_path.write_text(
        json.dumps(json_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    workbook = Workbook()
    article_sheet = workbook.active
    article_sheet.title = "articles"
    article_sheet.append(ARTICLES_COLUMNS)

    for record in records:
        article_sheet.append(
            [
                record["note_id"],
                record["title"],
                record["nickname"],
                record["note_url"],
                " ".join(f"#{tag}" for tag in record["tags"]),
                record["desc"],
                record["comments"][0],
                record["comments"][1],
                record["comments"][2],
                record["comments"][3],
                record["comments"][4],
                " / ".join(record["matched_topics"]),
            ]
        )

    comments_sheet = workbook.create_sheet("comments")
    comments_sheet.append(COMMENTS_COLUMNS)
    for record in records:
        for comment_index, comment_text in enumerate(record["comments"], start=1):
            comments_sheet.append(
                [
                    record["note_id"],
                    record["title"],
                    record["nickname"],
                    record["note_url"],
                    comment_index,
                    comment_text,
                ]
            )

    failed_sheet = workbook.create_sheet("failed")
    failed_sheet.append(FAILED_COLUMNS)
    for item in failed:
        failed_sheet.append([item[column] for column in FAILED_COLUMNS])

    workbook.save(xlsx_path)
    return {"json_path": str(json_path), "xlsx_path": str(xlsx_path)}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="根据小红书主页笔记生成短评 Excel")
    parser.add_argument("--profile-url", help="带 token 的小红书主页 URL")
    parser.add_argument("--input", help="已抓取的主页笔记 JSON；若未传则需要 --profile-url")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--limit", type=int, default=0, help="最多处理多少篇；默认 0 表示全部")
    parser.add_argument("--max-scroll-rounds", type=int, default=80, help="主页最多滚动轮数")
    parser.add_argument("--max-idle-scroll-rounds", type=int, default=4, help="主页连续无新增阈值")
    parser.add_argument("--delay-seconds", type=float, default=0.0, help="读取文章前的固定延迟秒数")
    parser.add_argument("--risk-stop-after", type=int, default=3, help="连续命中官方风控多少次后停止整批任务")
    parser.add_argument("--note-read-cooldown-min-seconds", type=float, default=4.0, help="相邻正文读取之间的最小冷却秒数")
    parser.add_argument("--note-read-cooldown-max-seconds", type=float, default=7.0, help="相邻正文读取之间的最大冷却秒数")
    parser.add_argument("--batch-read-cooldown-every", type=int, default=3, help="每读取多少篇正文后触发一次批次冷却")
    parser.add_argument("--batch-read-cooldown-min-seconds", type=float, default=20.0, help="正文抓取批次冷却最小秒数")
    parser.add_argument("--batch-read-cooldown-max-seconds", type=float, default=35.0, help="正文抓取批次冷却最大秒数")
    args = parser.parse_args()

    if not args.input and not args.profile_url:
        parser.error("必须提供 --input 或 --profile-url 其中之一")

    delay_seconds = resolve_delay_seconds(args.delay_seconds)
    notes = load_note_list(args.input)

    if args.profile_url:
        notes = fetch_notes_from_profile(
            args.profile_url,
            limit=(None if int(args.limit) <= 0 else int(args.limit)),
            max_scroll_rounds=int(args.max_scroll_rounds),
            max_idle_scroll_rounds=int(args.max_idle_scroll_rounds),
            delay_seconds=delay_seconds,
        )
    elif int(args.limit) > 0:
        notes = notes[: int(args.limit)]

    records, failed, summary = build_comment_records(
        notes,
        delay_seconds=delay_seconds,
        risk_stop_after=args.risk_stop_after,
        note_read_cooldown_min_seconds=args.note_read_cooldown_min_seconds,
        note_read_cooldown_max_seconds=args.note_read_cooldown_max_seconds,
        batch_read_cooldown_every=args.batch_read_cooldown_every,
        batch_read_cooldown_min_seconds=args.batch_read_cooldown_min_seconds,
        batch_read_cooldown_max_seconds=args.batch_read_cooldown_max_seconds,
    )
    paths = write_outputs(records, failed, summary, Path(args.output_dir).expanduser().resolve())
    print(
        json.dumps(
            {
                "count": len(records),
                "failed_count": len(failed),
                **summary,
                "output_dir": str(Path(args.output_dir).expanduser().resolve()),
                **paths,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
